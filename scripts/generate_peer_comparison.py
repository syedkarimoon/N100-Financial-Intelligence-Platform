"""
Sprint 3 - Day 20
Peer Comparison Excel Report

Generates:
    output/peer_comparison.xlsx

Workbook structure:
    - 11 sheets, one per peer group
    - company_id
    - company_name
    - 10 financial metrics
    - 10 percentile rank columns
    - benchmark company highlighted
    - peer-group median summary row
"""

import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# ============================================================
# PATHS
# ============================================================

DB_PATH = Path("db/nifty100.db")

PEER_GROUPS_PATH = Path(
    "data/supporting/peer_groups.xlsx"
)

OUTPUT_PATH = Path(
    "output/peer_comparison.xlsx"
)


# ============================================================
# METRIC MAPPING
# ============================================================

METRICS = [
    "ROE",
    "ROCE",
    "Net Profit Margin",
    "D/E",
    "FCF",
    "PAT CAGR 5yr",
    "Revenue CAGR 5yr",
    "EPS CAGR 5yr",
    "Interest Coverage",
    "Asset Turnover",
]


# ============================================================
# LOAD DATA
# ============================================================

def load_peer_percentiles(conn):
    """
    Load peer percentile results from SQLite.
    """

    query = """
    SELECT
        company_id,
        peer_group_name,
        metric,
        value,
        percentile_rank,
        year
    FROM peer_percentiles
    """

    return pd.read_sql_query(
        query,
        conn
    )


def load_company_names(conn):
    """
    Load company IDs and names.
    """

    query = """
    SELECT
        id AS company_id,
        company_name
    FROM companies
    """

    return pd.read_sql_query(
        query,
        conn
    )


def load_peer_groups():
    """
    Load peer group assignments and benchmark flags.
    """

    df = pd.read_excel(
        PEER_GROUPS_PATH
    )

    required_columns = {
        "company_id",
        "peer_group_name",
        "is_benchmark",
    }

    missing = (
        required_columns
        - set(df.columns)
    )

    if missing:
        raise ValueError(
            f"Missing columns in peer_groups.xlsx: "
            f"{missing}"
        )

    return df[
        [
            "company_id",
            "peer_group_name",
            "is_benchmark",
        ]
    ]


# ============================================================
# PREPARE REPORT DATA
# ============================================================

def prepare_report_data(
    peer_percentiles,
    company_names,
    peer_groups
):
    """
    Convert long-format percentile data
    into one row per company.

    Result:
        company_id
        company_name
        10 metric columns
        10 percentile columns
        peer_group_name
        is_benchmark
    """

    # --------------------------------------------------------
    # Pivot raw metric values
    # --------------------------------------------------------

    values = peer_percentiles.pivot_table(
        index=[
            "company_id",
            "peer_group_name",
        ],
        columns="metric",
        values="value",
        aggfunc="first",
    ).reset_index()

    # --------------------------------------------------------
    # Pivot percentile ranks
    # --------------------------------------------------------

    percentiles = peer_percentiles.pivot_table(
        index=[
            "company_id",
            "peer_group_name",
        ],
        columns="metric",
        values="percentile_rank",
        aggfunc="first",
    ).reset_index()

    # --------------------------------------------------------
    # Rename percentile columns
    # --------------------------------------------------------

    percentile_rename = {
        metric:
            f"{metric} Percentile"
        for metric in METRICS
    }

    percentiles = percentiles.rename(
        columns=percentile_rename
    )

    # --------------------------------------------------------
    # Merge values and percentiles
    # --------------------------------------------------------

    report = values.merge(
        percentiles,
        on=[
            "company_id",
            "peer_group_name",
        ],
        how="outer",
    )

    # --------------------------------------------------------
    # Add company names
    # --------------------------------------------------------

    report = report.merge(
        company_names,
        on="company_id",
        how="left",
    )

    # --------------------------------------------------------
    # Add benchmark flag
    # --------------------------------------------------------

    report = report.merge(
        peer_groups,
        on=[
            "company_id",
            "peer_group_name",
        ],
        how="left",
    )

    # --------------------------------------------------------
    # Keep required columns
    # --------------------------------------------------------

    columns = [
        "company_id",
        "company_name",
    ]

    for metric in METRICS:

        if metric in report.columns:

            columns.append(
                metric
            )

        percentile_column = (
            f"{metric} Percentile"
        )

        if percentile_column in report.columns:

            columns.append(
                percentile_column
            )

    report = report[
        [
            "peer_group_name",
            *columns,
            "is_benchmark",
        ]
    ]

    return report


# ============================================================
# WRITE EXCEL
# ============================================================

def write_excel(report):
    """
    Write one worksheet per peer group.
    """

    OUTPUT_PATH.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    # --------------------------------------------------------
    # Create workbook
    # --------------------------------------------------------

    with pd.ExcelWriter(
        OUTPUT_PATH,
        engine="openpyxl",
    ) as writer:

        peer_groups = sorted(
            report[
                "peer_group_name"
            ]
            .dropna()
            .unique()
        )

        for peer_group in peer_groups:

            sheet_data = report[
                report[
                    "peer_group_name"
                ]
                == peer_group
            ].copy()

            # Remove helper columns
            sheet_data = sheet_data.drop(
                columns=[
                    "peer_group_name",
                    "is_benchmark",
                ]
            )

            # --------------------------------------------------------
            # Add peer-group median summary row
            # --------------------------------------------------------

            median_row = {}

            for column in sheet_data.columns:

              if column in METRICS:

                median_row[column] = (
                sheet_data[column]
                .median()
               )

            else:

                median_row[column] = None


            # Label the summary row
            median_row["company_id"] = (
              "Peer Group Median"
            )

            median_row["company_name"] = (
              "Peer Group Median"
            )

            # Add summary row to bottom
            sheet_data = pd.concat(
             [
               sheet_data,
               pd.DataFrame(
               [median_row]
               ),
             ],
             ignore_index=True,
            )


            sheet_data.to_excel(
              writer,
              sheet_name=peer_group[
              :31
              ],
              index=False,
            )

    # --------------------------------------------------------
    # Apply formatting
    # --------------------------------------------------------

    workbook = load_workbook(
        OUTPUT_PATH
    )

    # Percentile fills
    green_fill = PatternFill(
        fill_type="solid",
        fgColor="C6EFCE",
    )

    yellow_fill = PatternFill(
        fill_type="solid",
        fgColor="FFEB9C",
    )

    red_fill = PatternFill(
        fill_type="solid",
        fgColor="FFC7CE",
    )

    # Benchmark fill
    benchmark_fill = PatternFill(
        fill_type="solid",
        fgColor="FFD966",
    )

    for worksheet in workbook.worksheets:

        # ----------------------------------------------------
        # Identify columns
        # ----------------------------------------------------

        headers = {
            cell.value: cell.column
            for cell in worksheet[1]
        }

        percentile_columns = [
            column
            for column in headers
            if (
                isinstance(
                    column,
                    str
                )
                and column.endswith(
                    " Percentile"
                )
            )
        ]

        # ----------------------------------------------------
        # Percentile colour coding
        # ----------------------------------------------------

        for row in worksheet.iter_rows(
            min_row=2,
            max_row=worksheet.max_row - 1
        ):

            for cell in row:

                if (
                    worksheet.cell(
                        1,
                        cell.column
                    ).value
                    in percentile_columns
                ):

                    if cell.value is None:
                        continue

                    if cell.value >= 0.75:

                        cell.fill = (
                            green_fill
                        )

                    elif cell.value > 0.25:

                        cell.fill = (
                            yellow_fill
                        )

                    else:

                        cell.fill = (
                            red_fill
                        )

        # ----------------------------------------------------
        # Benchmark highlighting
        # ----------------------------------------------------

        # Benchmark IDs are recovered
        # from the original report data.

        peer_group = worksheet.title

        benchmark_ids = report[
            (
                report[
                    "peer_group_name"
                ]
                == peer_group
            )
            &
            (
                report[
                    "is_benchmark"
                ]
                == True
            )
        ][
            "company_id"
        ].tolist()

        for row in worksheet.iter_rows(
            min_row=2
        ):

            company_id = row[0].value

            if company_id in benchmark_ids:

                for cell in row:

                    cell.fill = (
                        benchmark_fill
                    )
        # ----------------------------------------------------
        # Highlight peer-group median summary row
        # ----------------------------------------------------

        median_row_number = (
          worksheet.max_row
        )

        for cell in worksheet[   
            median_row_number
        ]:

            cell.fill = PatternFill(
               fill_type="solid",
               fgColor="D9EAD3",
            )
    # --------------------------------------------------------
    # Save formatted workbook
    # --------------------------------------------------------

    workbook.save(
        OUTPUT_PATH
    )

    print(
        "\nPeer comparison report generated:"
    )

    print(
        OUTPUT_PATH
    )

    print(
        "Sheets:",
        len(
            workbook.sheetnames
        )
    )

    print(
        "Sheet names:",
        workbook.sheetnames
    )


# ============================================================
# MAIN
# ============================================================

def main():

    print(
        "\n"
        + "=" * 70
    )

    print(
        "SPRINT 3 - DAY 20"
    )

    print(
        "PEER COMPARISON EXCEL REPORT"
    )

    print(
        "=" * 70
    )

    # --------------------------------------------------------
    # Database connection
    # --------------------------------------------------------

    conn = sqlite3.connect(
        DB_PATH
    )

    try:

        # ----------------------------------------------------
        # Load data
        # ----------------------------------------------------

        print(
            "\nLoading peer percentile data..."
        )

        peer_percentiles = (
            load_peer_percentiles(
                conn
            )
        )

        print(
            "Percentile records:",
            len(
                peer_percentiles
            )
        )

        print(
            "\nLoading company names..."
        )

        company_names = (
            load_company_names(
                conn
            )
        )

        print(
            "Companies:",
            len(
                company_names
            )
        )

        print(
            "\nLoading peer groups..."
        )

        peer_groups = (
            load_peer_groups()
        )

        print(
            "Peer group records:",
            len(
                peer_groups
            )
        )

        # ----------------------------------------------------
        # Prepare report
        # ----------------------------------------------------

        print(
            "\nPreparing report data..."
        )

        report = prepare_report_data(
            peer_percentiles,
            company_names,
            peer_groups,
        )

        print(
            "Report rows:",
            len(
                report
            )
        )

        print(
            "Peer groups:",
            report[
                "peer_group_name"
            ]
            .nunique()
        )

        # ----------------------------------------------------
        # Write Excel
        # ----------------------------------------------------

        print(
            "\nWriting Excel report..."
        )

        write_excel(
            report
        )

    finally:

        conn.close()


if __name__ == "__main__":
    main()